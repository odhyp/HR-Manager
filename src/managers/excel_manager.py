from openpyxl import load_workbook, utils
from openpyxl.styles import (Border,
                             Color,
                             Font,
                             NamedStyle,
                             PatternFill,
                             Side)
from win32com.client import Dispatch
from xls2xlsx import XLS2XLSX


class ExcelManager:
    def convert_xls(self, file_path: str, file_format: int = 51):
        """Convert an Excel file from '.xls' to '.xlsx' format using pywin32

        Args:
        - file_path (str): The path to the input '.xls' file.
        - file_format (int, optional): The file format code for '.xlsx'.
        Default is 51.
        """
        excel_app = Dispatch("Excel.Application")
        excel_app.Visible = False
        output_path = str(file_path) + 'x'
        workbook = excel_app.Workbooks.Open(file_path)
        workbook.SaveAs(output_path, FileFormat=file_format)
        workbook.Close()
        excel_app.Quit()

    def convert_xls2(self, file_path: str):
        """Convert an Excel file from '.xls' to '.xlsx' format using xls2xlsx.

        Args:
        - file_path (str): The path to the input '.xls' file.
        """
        output_path = str(file_path) + 'x'
        x2x = XLS2XLSX(file_path)
        x2x.to_xlsx(output_path)

    def format_nominatif(self, file_path: str):
        """Format 'Nominatif.xlsx'

        Args:
        - file_path (str): Path to the 'Nominatif.xlsx'
        """
        wb = load_workbook(filename=file_path)
        ws = wb.active
        ws.delete_cols(1, 1)
        ws.delete_rows(1, 4)
        wb.save(filename=file_path)

    def format_duk(self, file_path: str):
        """Format 'DUK.xlsx'

        Args:
        - file_path (str): Path to the 'DUK.xlsx'
        """
        wb = load_workbook(filename=file_path)
        ws = wb.active
        ws.delete_rows(1, 4)
        wb.save(filename=file_path)

    def format_presensi(self, file_path: str):
        """Format 'Presensi.xlsx'

        Args:
        - file_path (str): Path to the 'Presensi.xlsx'
        """
        wb = load_workbook(filename=file_path)
        ws = wb.active

        # Define last row, last column, and cell range
        last_row = ws.max_row
        last_column = ws.max_column
        cell_range = f"A1:{utils.get_column_letter(last_column)}{last_row}"

        # Set font colors to black
        for row in ws[cell_range]:
            for cell in row:
                cell.font = Font(color='000000', bold=False)

        # Create styles for conditional formatting
        cf_red = NamedStyle(
            name='cf_red',
            font=Font(color='9C0006'),
            fill=PatternFill(start_color='FFC7CE', fill_type='solid')
        )

        cf_yellow = NamedStyle(
            name='cf_yellow',
            font=Font(color='9C6500'),
            fill=PatternFill(start_color='FFEB9C', fill_type='solid')
        )

        cf_green = NamedStyle(
            name='cf_green',
            font=Font(color='006100'),
            fill=PatternFill(start_color='C6EFCE', fill_type='solid')
        )

        # Apply conditional formatting rule
        for column in ws.iter_cols(min_col=4,
                                   max_col=last_column,
                                   min_row=2,
                                   max_row=last_row):
            for cell in column:
                if cell.value is not None:
                    # Bad
                    if "TK" in str(cell.value):
                        cell.style = cf_red
                    elif "TPD" in str(cell.value):
                        cell.style = cf_red
                    elif "TPP" in str(cell.value):
                        cell.style = cf_red

                    # Yellow
                    elif "CS" in str(cell.value):
                        cell.style = cf_yellow
                    elif "CT" in str(cell.value):
                        cell.style = cf_yellow
                    elif "CAP" in str(cell.value):
                        cell.style = cf_yellow
                    elif "CBS" in str(cell.value):
                        cell.style = cf_yellow
                    elif "CB" in str(cell.value):
                        cell.style = cf_yellow

                    # Green
                    elif "TD" in str(cell.value):
                        cell.style = cf_green

        # Add borders
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Hide column 'B'
        ws.column_dimensions['B'].hidden = True

        wb.save(filename=file_path)

    def format_prestasi(self, file_path: str):
        """Format 'Prestasi.xlsx'

        Args:
        - file_path (str): Path to the 'Presensi.xlsx'
        """
        wb = load_workbook(filename=file_path)
        ws = wb.active
        ws.column_dimensions['D'].hidden = True
        wb.save(filename=file_path)

    def format_xlsx(self, file_type: str, file_path: str):
        """Format the specified 'xlsx' file based on the file type. This method
        wraps 'formatting' methods for 'xlsx' file.

        Args:
        - file_type (str): The type of the file
        - file_path (str): Path to the 'xlsx' file
        """
        if file_type == 'nominatif':
            self.format_nominatif(file_path=file_path)
        elif file_type == 'duk':
            self.format_duk(file_path=file_path)
        elif file_type == 'presensi':
            self.format_presensi(file_path=file_path)
        elif file_type == 'prestasi':
            self.format_prestasi(file_path=file_path)
        else:
            return f"Invalid type: {file_type} is not supported."
